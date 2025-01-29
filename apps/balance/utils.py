import openpyxl
import io
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .models import RegistroPeso


def generar_excel_registro_peso():
    """
    Generar un excel que contenga los registros de peso.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = f"Registro de Pesajes"

    encabezados = [
        "# Parte",
        "Fecha de Emisión",
        "Tipo de Cliente",
        "Cod. Almacen",
        "Desc. Almacen",
        "Documento",
        "Entidad Origen",
        "Conductor",
        "Placa",
        "Estado",
        "Hecho por",
        "Fecha",
        "Hora",
    ]

    # Definir estilos
    header_font = Font(bold=True, color="000000", size=12) # letra
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # alineación
    
    # Definir los estilos de los bordes
    border_style = Side(style='thin', color='000000')
    header_border = Border(
        bottom=border_style,
    )

    ws.row_dimensions[2].height = 50 # altura header fijo

    # Encabezados
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=2, column=col, value=encabezado)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    ws.cell(row=1, column=1, value="Registro de Pesajes")

    registros_pesos = RegistroPeso.objects.all()

    fila = 3
    for i, registro in enumerate(registros_pesos, start=1):
        ws.cell(row=fila, column=1, value=registro.id)
        ws.cell(row=fila, column=2, value=registro.fecha)
        ws.cell(row=fila, column=3, value=registro.tipo_cliente)
        ws.cell(row=fila, column=4, value=registro.sede.codigo)
        ws.cell(row=fila, column=5, value=registro.almacen.name)
        ws.cell(row=fila, column=6, value=registro.documento_numero)
        ws.cell(row=fila, column=7, value=registro.entidad_origen)
        ws.cell(row=fila, column=8, value=registro.conductor)
        ws.cell(row=fila, column=9, value=registro.placa_camion)
        ws.cell(row=fila, column=12, value=registro.fecha_ingreso_balanza)
        ws.cell(row=fila, column=13, value=registro.hora_ingreso_balanza)

        fila += 1

    # Encabezados autoajustables
    for col in range(1, len(encabezados)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_size = True

    # Guardamos en un buffer de memoria en lugar de un archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()